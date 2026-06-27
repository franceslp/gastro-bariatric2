#!/usr/bin/env python3
"""
build_masterfile.py

Builds the complete updated masterfile Excel workbook with all
correct funnel steps from the current validated pipeline.

Run on the VM in ~/bariatric_analysis/
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "masterfile_gastro_bariatric2.xlsx"

# Header style: dark blue background, white bold text
HDR_FILL = PatternFill("solid", start_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Alternating row fills
ROW_FILL_A = PatternFill("solid", start_color="FFFFFF")
ROW_FILL_B = PatternFill("solid", start_color="EBF3FB")

# Highlight fill for excluded/special rows
EXCL_FILL = PatternFill("solid", start_color="FFE0E0")
CONSORT_FILL = PatternFill("solid", start_color="E2EFDA")

DATA_FONT = Font(name="Arial", size=9)
DATA_ALIGN = Alignment(horizontal="left", vertical="center")

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_sheet(wb, sheet_name, df, title=None, highlight_col=None, highlight_fill=None):
    ws = wb.create_sheet(sheet_name)
    row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, name="Arial", size=11)
        ws.cell(row=1, column=1).fill = PatternFill("solid", start_color="D6E4F0")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        row = 2

    # Header
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=row, column=c, value=col)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = BORDER
    row += 1

    # Data rows
    for i, (_, data_row) in enumerate(df.iterrows()):
        fill = ROW_FILL_B if i % 2 == 0 else ROW_FILL_A
        if highlight_col and highlight_col in df.columns:
            val = str(data_row.get(highlight_col, ""))
            if val and val != "nan":
                fill = highlight_fill or fill
        for c, val in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=c, value=val if str(val) != "nan" else "")
            cell.font = DATA_FONT
            cell.fill = fill
            cell.alignment = DATA_ALIGN
            cell.border = BORDER
        row += 1

    # Auto-width columns
    for c, col in enumerate(df.columns, 1):
        max_len = max(len(str(col)), 10)
        for r in ws.iter_rows(min_row=2, max_row=min(row, 52), min_col=c, max_col=c):
            for cell in r:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 40))
        ws.column_dimensions[get_column_letter(c)].width = max_len + 2

    ws.freeze_panes = ws.cell(row=(3 if title else 2), column=1)
    return ws

print("Reading pipeline files...")

# Load all funnel files
f1  = pd.read_csv("funnel_1_all_patients_1118.csv",  dtype=str)
f2  = pd.read_csv("funnel_step1_age.csv",             dtype=str)
f3  = pd.read_csv("funnel_step2_k3184_1yr.csv",       dtype=str)
f4  = pd.read_csv("funnel_step3_e10e11_1yr.csv",      dtype=str)
f5  = pd.read_csv("funnel_step4_ges.csv",             dtype=str)
f6  = pd.read_csv("cohort_FINAL_analytic.csv",        dtype=str)
excl = pd.read_csv("excluded_multisurgery_sameday.csv", dtype=str)
comp = pd.read_csv("comparator_pool_raw.csv",         dtype=str)

print(f"  Step 1 (all):        {len(f1):,}")
print(f"  Step 2 (age ≥18):    {len(f2):,}")
print(f"  Step 3 (K31.84 1yr): {len(f3):,}")
print(f"  Step 4 (E10/E11):    {len(f4):,}")
print(f"  Step 5 (GES):        {len(f5):,}")
print(f"  Step 6 (final):      {len(f6):,}")
print(f"  Excluded step 6:     {len(excl):,}")
print(f"  Comparator pool:     {len(comp):,}")

# CONSORT summary sheet
consort_data = {
    "Step": [
        "Starting pool (GP + DM + Bariatric surgery, 2015–2025)",
        "Exclusion 1: Age < 18 at surgery",
        "After age exclusion",
        "Exclusion 2: No K31.84 within 1 year before surgery",
        "After K31.84 timing exclusion",
        "Exclusion 3: No E10/E11 within 1 year before surgery",
        "After diabetes exclusion",
        "Exclusion 4: No GES before K31.84 diagnosis",
        "After GES exclusion",
        "Exclusion 5: Ambiguous/multiple bariatric surgery records",
        "  → Same-day conflicting CPT codes",
        "  → Surgeries >180 days apart (revision/conversion)",
        "  → Surgeries 0-2 days apart (billing artifact)",
        "  → Surgeries weeks-to-months apart",
        "FINAL GP ANALYTIC COHORT",
        "",
        "Comparator pool (bariatric surgery, never K31.84, single procedure)",
    ],
    "n": [
        1118, -1, 1117, -210, 907, -28, 879, -495, 384,
        -8, -3, -2, -2, -1, 376,
        "", 7027
    ],
    "Notes": [
        "ICD-10 K31.84 + E10/E11 + CPT 43644/43645/43775/43846/43847",
        "", "",
        "K31.84 must occur within 365 days before bariatric surgery", "", "",
        "E10 or E11 must occur within 365 days before surgery", "", "",
        "GES (CPT 43647/43881/43882) must precede K31.84 diagnosis",
        "Sleeve + bypass CPT on same date — index unclear",
        "Second bariatric surgery >180d after first",
        "Second bariatric surgery 0-2d — likely split billing",
        "Second bariatric surgery weeks-months — likely staged procedure",
        "GP-bariatric patients with confirmed gastroparesis and GES",
        "",
        "All bariatric surgery patients without K31.84, single clean procedure"
    ]
}
consort_df = pd.DataFrame(consort_data)
consort_df["n"] = consort_df["n"].astype(str).replace("-1","").replace("","")
# Clean up negatives display
def fmt_n(v):
    try:
        iv = int(v)
        return f"−{abs(iv)}" if iv < 0 else str(iv)
    except:
        return str(v)
consort_df["n"] = consort_df["n"].apply(fmt_n)

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# Sheet 0: CONSORT
ws0 = wb.create_sheet("CONSORT Flow")
ws0.sheet_tab_color = "1F4E79"
ws0.cell(1,1,"CONSORT Flow Diagram — Gastro-Bariatric2 Study").font = Font(bold=True, name="Arial", size=12)
ws0.cell(1,1).fill = PatternFill("solid", start_color="1F4E79")
ws0.cell(1,1).font = Font(bold=True, name="Arial", size=12, color="FFFFFF")
ws0.merge_cells("A1:C1")
hdrs = ["Step", "n", "Notes"]
for c, h in enumerate(hdrs, 1):
    cell = ws0.cell(2, c, h)
    cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = HDR_ALIGN; cell.border = BORDER
for i, row in consort_df.iterrows():
    is_final = "FINAL" in str(row["Step"])
    is_excl = str(row["n"]).startswith("−")
    fill = PatternFill("solid", start_color="E2EFDA") if is_final else (
           PatternFill("solid", start_color="FFE8E8") if is_excl else
           (ROW_FILL_B if i%2==0 else ROW_FILL_A))
    for c, val in enumerate([row["Step"], row["n"], row["Notes"]], 1):
        cell = ws0.cell(i+3, c, val)
        cell.font = Font(bold=is_final, name="Arial", size=9)
        cell.fill = fill; cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")
ws0.column_dimensions["A"].width = 60
ws0.column_dimensions["B"].width = 10
ws0.column_dimensions["C"].width = 55
ws0.freeze_panes = "A3"

# Remaining sheets
sheets_to_write = [
    ("1- All (n=1,118)",        f1,   "Step 1: All patients with GP + Diabetes + Bariatric Surgery (n=1,118)"),
    ("2- Age ≥18 (n=1,117)",    f2,   "Step 2: After age ≥18 exclusion (n=1,117)"),
    ("3- K31.84 1yr (n=907)",   f3,   "Step 3: After K31.84 within 1yr requirement (n=907)"),
    ("4- E10E11 1yr (n=879)",   f4,   "Step 4: After E10/E11 diabetes within 1yr requirement (n=879)"),
    ("5- GES (n=384)",          f5,   "Step 5: After GES before K31.84 requirement (n=384)"),
    ("6- Excluded step5 (n=8)", excl, "Step 5 exclusions: Ambiguous/multiple bariatric surgery records (n=8)"),
    ("7- Final GP (n=376)",     f6,   "FINAL GP Analytic Cohort (n=376)"),
    ("8- Comparator (n=7,027)", comp, "Comparator Pool: Bariatric surgery patients without gastroparesis (n=7,027)"),
]

tab_colors = ["2E75B6","2E75B6","2E75B6","2E75B6","2E75B6","FF0000","70AD47","ED7D31"]

for (sname, df, title), color in zip(sheets_to_write, tab_colors):
    ws = write_sheet(wb, sname, df, title=title,
                     highlight_col="exclusion_reason" if "Excluded" in sname else None,
                     highlight_fill=EXCL_FILL)
    ws.sheet_tab_color = color

wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")
print("Sheet summary:")
for ws in wb.worksheets:
    print(f"  {ws.title}")
